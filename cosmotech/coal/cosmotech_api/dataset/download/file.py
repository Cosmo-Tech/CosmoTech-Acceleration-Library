# Copyright (C) - 2023 - 2025 - Cosmo Tech
# This document and all information contained herein is the exclusive property -
# including all intellectual property rights pertaining thereto - of Cosmo Tech.
# Any use, reproduction, translation, broadcasting, transmission, distribution,
# etc., to any person is prohibited unless it has been previously and
# specifically authorized by written means by Cosmo Tech.

import csv
import io
import json
import os
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, Tuple

from cosmotech_api import WorkspaceApi
from openpyxl import load_workbook

from cosmotech.coal.utils.logger import LOGGER
from cosmotech.orchestrator.utils.translate import T
from cosmotech.coal.cosmotech_api.connection import get_api_client


def download_file_dataset(
    organization_id: str,
    workspace_id: str,
    file_name: str,
    target_folder: Optional[Union[str, Path]] = None,
    read_files: bool = True,
) -> Tuple[Dict[str, Any], Path]:
    """
    Download file dataset.

    Args:
        organization_id: Organization ID
        workspace_id: Workspace ID
        file_name: File name to download
        target_folder: Optional folder to save files (if None, uses temp dir)
        read_files: Whether to read file contents

    Returns:
        Tuple of (content dict, folder path)
    """
    start_time = time.time()
    LOGGER.info(T("coal.services.dataset.download_started").format(dataset_type="File"))
    LOGGER.debug(
        T("coal.services.dataset.file_downloading").format(
            organization_id=organization_id,
            workspace_id=workspace_id,
            file_name=file_name,
        )
    )

    # Create temp directory for downloaded files
    if target_folder is None:
        tmp_dataset_dir = tempfile.mkdtemp()
    else:
        tmp_dataset_dir = Path(target_folder)
        tmp_dataset_dir.mkdir(parents=True, exist_ok=True)
        tmp_dataset_dir = str(tmp_dataset_dir)

    LOGGER.debug(T("coal.services.dataset.using_folder").format(folder=tmp_dataset_dir))

    content = dict()

    with get_api_client()[0] as api_client:
        api_ws = WorkspaceApi(api_client)

        # Find all files matching the pattern
        list_start = time.time()
        LOGGER.debug(T("coal.services.dataset.listing_workspace_files"))
        all_api_files = api_ws.find_all_workspace_files(organization_id, workspace_id)

        existing_files = list(_f.file_name for _f in all_api_files if _f.file_name.startswith(file_name))
        list_time = time.time() - list_start

        LOGGER.debug(T("coal.services.dataset.workspace_files_found").format(count=len(existing_files)))
        LOGGER.debug(T("coal.common.timing.operation_completed").format(operation="file listing", time=list_time))

        if not existing_files:
            LOGGER.warning(T("coal.services.dataset.no_files_found").format(file_name=file_name))
            return content, Path(tmp_dataset_dir)

        # Download and process each file
        for _file_name in existing_files:
            download_start = time.time()
            LOGGER.debug(T("coal.services.dataset.downloading_file").format(file_name=_file_name))

            dl_file = api_ws.download_workspace_file(
                organization_id=organization_id,
                workspace_id=workspace_id,
                file_name=_file_name,
            )

            target_file = os.path.join(tmp_dataset_dir, _file_name.split("/")[-1])
            with open(target_file, "wb") as tmp_file:
                tmp_file.write(dl_file)

            download_time = time.time() - download_start
            LOGGER.debug(T("coal.services.dataset.file_downloaded").format(file_name=_file_name, path=target_file))
            LOGGER.debug(
                T("coal.common.timing.operation_completed").format(
                    operation=f"download {_file_name}", time=download_time
                )
            )

            if not read_files:
                continue

            # Process file based on type
            process_start = time.time()

            if ".xls" in _file_name:
                LOGGER.debug(T("coal.services.dataset.processing_excel").format(file_name=target_file))
                wb = load_workbook(target_file, data_only=True)

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    content[sheet_name] = list()
                    headers = next(sheet.iter_rows(max_row=1, values_only=True))

                    def item(_row: tuple) -> dict:
                        return {k: v for k, v in zip(headers, _row)}

                    row_count = 0
                    for r in sheet.iter_rows(min_row=2, values_only=True):
                        row = item(r)
                        new_row = dict()

                        for key, value in row.items():
                            try:
                                converted_value = json.load(io.StringIO(value))
                            except (json.decoder.JSONDecodeError, TypeError):
                                converted_value = value

                            if converted_value is not None:
                                new_row[key] = converted_value

                        if new_row:
                            content[sheet_name].append(new_row)
                            row_count += 1

                    LOGGER.debug(
                        T("coal.services.dataset.sheet_processed").format(sheet_name=sheet_name, rows=row_count)
                    )

            elif ".csv" in _file_name:
                LOGGER.debug(T("coal.services.dataset.processing_csv").format(file_name=target_file))
                with open(target_file, "r") as file:
                    current_filename = os.path.basename(target_file)[: -len(".csv")]
                    content[current_filename] = list()

                    row_count = 0
                    for csv_row in csv.DictReader(file):
                        csv_row: dict
                        new_row = dict()

                        for key, value in csv_row.items():
                            try:
                                # Try to convert any json row to dict object
                                converted_value = json.load(io.StringIO(value))
                            except json.decoder.JSONDecodeError:
                                converted_value = value

                            if converted_value == "":
                                converted_value = None

                            if converted_value is not None:
                                new_row[key] = converted_value

                        content[current_filename].append(new_row)
                        row_count += 1

                    LOGGER.debug(
                        T("coal.services.dataset.csv_processed").format(file_name=current_filename, rows=row_count)
                    )

            elif ".json" in _file_name:
                LOGGER.debug(T("coal.services.dataset.processing_json").format(file_name=target_file))
                with open(target_file, "r") as _file:
                    current_filename = os.path.basename(target_file)
                    content[current_filename] = json.load(_file)

                    if isinstance(content[current_filename], dict):
                        item_count = len(content[current_filename])
                    elif isinstance(content[current_filename], list):
                        item_count = len(content[current_filename])
                    else:
                        item_count = 1

                    LOGGER.debug(
                        T("coal.services.dataset.json_processed").format(file_name=current_filename, items=item_count)
                    )

            else:
                LOGGER.debug(T("coal.services.dataset.processing_text").format(file_name=target_file))
                with open(target_file, "r") as _file:
                    current_filename = os.path.basename(target_file)
                    content[current_filename] = "\n".join(line for line in _file)

                    line_count = content[current_filename].count("\n") + 1
                    LOGGER.debug(
                        T("coal.services.dataset.text_processed").format(file_name=current_filename, lines=line_count)
                    )

            process_time = time.time() - process_start
            LOGGER.debug(
                T("coal.common.timing.operation_completed").format(operation=f"process {_file_name}", time=process_time)
            )

    elapsed_time = time.time() - start_time
    LOGGER.info(T("coal.common.timing.operation_completed").format(operation="File download", time=elapsed_time))
    LOGGER.info(T("coal.services.dataset.download_completed").format(dataset_type="File"))

    return content, Path(tmp_dataset_dir)
