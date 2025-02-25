# Copyright (C) - 2023 - 2025 - Cosmo Tech
# This document and all information contained herein is the exclusive property -
# including all intellectual property rights pertaining thereto - of Cosmo Tech.
# Any use, reproduction, translation, broadcasting, transmission, distribution,
# etc., to any person is prohibited unless it has been previously and
# specifically authorized by written means by Cosmo Tech.

import csv
import io
import json
import os
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, Tuple

from azure.digitaltwins.core import DigitalTwinsClient
from azure.identity import DefaultAzureCredential
from cosmotech_api import DatasetApi, DatasetTwinGraphQuery, TwinGraphQuery, TwingraphApi, WorkspaceApi
from openpyxl import load_workbook

from cosmotech.coal.utils.logger import LOGGER
from cosmotech.orchestrator.utils.translate import T
from cosmotech.coal.cosmotech_api.connection import get_api_client
from cosmotech.coal.dataset.utils import get_content_from_twin_graph_data
from cosmotech.coal.dataset.converters import convert_dataset_to_files


def download_adt_dataset(
    adt_address: str,
    target_folder: Optional[Union[str, Path]] = None,
    credentials: Optional[DefaultAzureCredential] = None
) -> Tuple[Dict[str, Any], Path]:
    """
    Download dataset from Azure Digital Twins.
    
    Args:
        adt_address: The ADT instance address
        target_folder: Optional folder to save files (if None, uses temp dir)
        credentials: Optional Azure credentials (if None, uses DefaultAzureCredential)
        
    Returns:
        Tuple of (content dict, folder path)
    """
    start_time = time.time()
    LOGGER.info(T("coal.logs.dataset.download_started").format(dataset_type="ADT"))
    LOGGER.debug(T("coal.logs.dataset.adt_connecting").format(url=adt_address))
    
    # Create credentials if not provided
    if credentials is None:
        if get_api_client()[1] == "Azure Entra Connection":
            credentials = DefaultAzureCredential()
        else:
            LOGGER.error(T("coal.logs.dataset.adt_no_credentials"))
            raise ValueError("No credentials available for ADT connection")
    
    # Create client and download data
    client = DigitalTwinsClient(adt_address, credentials)
    
    # Query twins
    query_start = time.time()
    LOGGER.debug(T("coal.logs.dataset.adt_querying_twins"))
    query_expression = 'SELECT * FROM digitaltwins'
    query_result = client.query_twins(query_expression)
    
    json_content = dict()
    twin_count = 0
    
    for twin in query_result:
        twin_count += 1
        entity_type = twin.get('$metadata').get('$model').split(':')[-1].split(';')[0]
        t_content = {k: v for k, v in twin.items()}
        t_content['id'] = t_content['$dtId']
        
        # Remove system properties
        for k in list(twin.keys()):
            if k[0] == '$':
                del t_content[k]
                
        json_content.setdefault(entity_type, [])
        json_content[entity_type].append(t_content)
    
    query_time = time.time() - query_start
    LOGGER.debug(T("coal.logs.dataset.adt_twins_found").format(count=twin_count))
    LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
        operation="twins query",
        time=query_time
    ))
    
    # Query relationships
    rel_start = time.time()
    LOGGER.debug(T("coal.logs.dataset.adt_querying_relations"))
    relations_query = 'SELECT * FROM relationships'
    query_result = client.query_twins(relations_query)
    
    relation_count = 0
    for relation in query_result:
        relation_count += 1
        tr = {
            "$relationshipId": "id",
            "$sourceId": "source",
            "$targetId": "target"
        }
        r_content = {k: v for k, v in relation.items()}
        
        # Map system properties to standard names
        for k, v in tr.items():
            r_content[v] = r_content[k]
            
        # Remove system properties
        for k in list(relation.keys()):
            if k[0] == '$':
                del r_content[k]
                
        json_content.setdefault(relation['$relationshipName'], [])
        json_content[relation['$relationshipName']].append(r_content)
    
    rel_time = time.time() - rel_start
    LOGGER.debug(T("coal.logs.dataset.adt_relations_found").format(count=relation_count))
    LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
        operation="relations query",
        time=rel_time
    ))
    
    # Convert to files if target_folder is provided
    if target_folder:
        dataset_info = {
            "type": "adt",
            "content": json_content,
            "name": "ADT Dataset"
        }
        target_folder = convert_dataset_to_files(dataset_info, target_folder)
    else:
        target_folder = tempfile.mkdtemp()
    
    elapsed_time = time.time() - start_time
    LOGGER.info(T("coal.logs.dataset.operation_timing").format(
        operation="ADT download",
        time=elapsed_time
    ))
    LOGGER.info(T("coal.logs.dataset.download_completed").format(dataset_type="ADT"))
    
    return json_content, Path(target_folder)


def download_twingraph_dataset(
    organization_id: str,
    dataset_id: str,
    target_folder: Optional[Union[str, Path]] = None
) -> Tuple[Dict[str, Any], Path]:
    """
    Download dataset from TwinGraph.
    
    Args:
        organization_id: Organization ID
        dataset_id: Dataset ID
        target_folder: Optional folder to save files (if None, uses temp dir)
        
    Returns:
        Tuple of (content dict, folder path)
    """
    start_time = time.time()
    LOGGER.info(T("coal.logs.dataset.download_started").format(dataset_type="TwinGraph"))
    LOGGER.debug(T("coal.logs.dataset.twingraph_downloading").format(
        organization_id=organization_id,
        dataset_id=dataset_id
    ))
    
    with get_api_client()[0] as api_client:
        dataset_api = DatasetApi(api_client)
        
        # Query nodes
        nodes_start = time.time()
        LOGGER.debug(T("coal.logs.dataset.twingraph_querying_nodes").format(dataset_id=dataset_id))
        nodes_query = DatasetTwinGraphQuery(query="MATCH(n) RETURN n")
        
        nodes = dataset_api.twingraph_query(
            organization_id=organization_id,
            dataset_id=dataset_id,
            dataset_twin_graph_query=nodes_query
        )
        
        nodes_time = time.time() - nodes_start
        LOGGER.debug(T("coal.logs.dataset.twingraph_nodes_found").format(count=len(nodes)))
        LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
            operation="nodes query",
            time=nodes_time
        ))
        
        # Query edges
        edges_start = time.time()
        LOGGER.debug(T("coal.logs.dataset.twingraph_querying_edges").format(dataset_id=dataset_id))
        edges_query = DatasetTwinGraphQuery(query="MATCH(n)-[r]->(m) RETURN n as src, r as rel, m as dest")
        
        edges = dataset_api.twingraph_query(
            organization_id=organization_id,
            dataset_id=dataset_id,
            dataset_twin_graph_query=edges_query
        )
        
        edges_time = time.time() - edges_start
        LOGGER.debug(T("coal.logs.dataset.twingraph_edges_found").format(count=len(edges)))
        LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
            operation="edges query",
            time=edges_time
        ))
        
        # Process results
        process_start = time.time()
        content = get_content_from_twin_graph_data(nodes, edges, True)
        process_time = time.time() - process_start
        
        LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
            operation="data processing",
            time=process_time
        ))
    
    # Convert to files if target_folder is provided
    if target_folder:
        dataset_info = {
            "type": "twincache",
            "content": content,
            "name": f"TwinGraph Dataset {dataset_id}"
        }
        target_folder = convert_dataset_to_files(dataset_info, target_folder)
    else:
        target_folder = tempfile.mkdtemp()
    
    elapsed_time = time.time() - start_time
    LOGGER.info(T("coal.logs.dataset.operation_timing").format(
        operation="TwinGraph download",
        time=elapsed_time
    ))
    LOGGER.info(T("coal.logs.dataset.download_completed").format(dataset_type="TwinGraph"))
    
    return content, Path(target_folder)


def download_legacy_twingraph_dataset(
    organization_id: str,
    cache_name: str,
    target_folder: Optional[Union[str, Path]] = None
) -> Tuple[Dict[str, Any], Path]:
    """
    Download dataset from legacy TwinGraph.
    
    Args:
        organization_id: Organization ID
        cache_name: Twin cache name
        target_folder: Optional folder to save files (if None, uses temp dir)
        
    Returns:
        Tuple of (content dict, folder path)
    """
    start_time = time.time()
    LOGGER.info(T("coal.logs.dataset.download_started").format(dataset_type="Legacy TwinGraph"))
    LOGGER.debug(T("coal.logs.dataset.legacy_twingraph_downloading").format(
        organization_id=organization_id,
        cache_name=cache_name
    ))
    
    with get_api_client()[0] as api_client:
        api_instance = TwingraphApi(api_client)
        
        # Query nodes
        nodes_start = time.time()
        LOGGER.debug(T("coal.logs.dataset.legacy_twingraph_querying_nodes").format(cache_name=cache_name))
        _query_nodes = TwinGraphQuery(query="MATCH(n) RETURN n")
        
        nodes = api_instance.query(
            organization_id=organization_id,
            graph_id=cache_name,
            twin_graph_query=_query_nodes
        )
        
        nodes_time = time.time() - nodes_start
        LOGGER.debug(T("coal.logs.dataset.legacy_twingraph_nodes_found").format(count=len(nodes)))
        LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
            operation="nodes query",
            time=nodes_time
        ))
        
        # Query relationships
        rel_start = time.time()
        LOGGER.debug(T("coal.logs.dataset.legacy_twingraph_querying_relations").format(cache_name=cache_name))
        _query_rel = TwinGraphQuery(query="MATCH(n)-[r]->(m) RETURN n as src, r as rel, m as dest")
        
        rel = api_instance.query(
            organization_id=organization_id,
            graph_id=cache_name,
            twin_graph_query=_query_rel
        )
        
        rel_time = time.time() - rel_start
        LOGGER.debug(T("coal.logs.dataset.legacy_twingraph_relations_found").format(count=len(rel)))
        LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
            operation="relations query",
            time=rel_time
        ))
        
        # Process results
        process_start = time.time()
        content = get_content_from_twin_graph_data(nodes, rel, False)
        process_time = time.time() - process_start
        
        LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
            operation="data processing",
            time=process_time
        ))
    
    # Convert to files if target_folder is provided
    if target_folder:
        dataset_info = {
            "type": "twincache",
            "content": content,
            "name": f"Legacy TwinGraph Dataset {cache_name}"
        }
        target_folder = convert_dataset_to_files(dataset_info, target_folder)
    else:
        target_folder = tempfile.mkdtemp()
    
    elapsed_time = time.time() - start_time
    LOGGER.info(T("coal.logs.dataset.operation_timing").format(
        operation="Legacy TwinGraph download",
        time=elapsed_time
    ))
    LOGGER.info(T("coal.logs.dataset.download_completed").format(dataset_type="Legacy TwinGraph"))
    
    return content, Path(target_folder)


def download_file_dataset(
    organization_id: str,
    workspace_id: str,
    file_name: str,
    target_folder: Optional[Union[str, Path]] = None,
    read_files: bool = True
) -> Tuple[Dict[str, Any], Path]:
    """
    Download file dataset.
    
    Args:
        organization_id: Organization ID
        workspace_id: Workspace ID
        file_name: File name to download
        target_folder: Optional folder to save files (if None, uses temp dir)
        read_files: Whether to read file contents
        
    Returns:
        Tuple of (content dict, folder path)
    """
    start_time = time.time()
    LOGGER.info(T("coal.logs.dataset.download_started").format(dataset_type="File"))
    LOGGER.debug(T("coal.logs.dataset.file_downloading").format(
        organization_id=organization_id,
        workspace_id=workspace_id,
        file_name=file_name
    ))
    
    # Create temp directory for downloaded files
    if target_folder is None:
        tmp_dataset_dir = tempfile.mkdtemp()
    else:
        tmp_dataset_dir = Path(target_folder)
        tmp_dataset_dir.mkdir(parents=True, exist_ok=True)
        tmp_dataset_dir = str(tmp_dataset_dir)
    
    LOGGER.debug(T("coal.logs.dataset.using_folder").format(folder=tmp_dataset_dir))
    
    content = dict()
    
    with get_api_client()[0] as api_client:
        api_ws = WorkspaceApi(api_client)
        
        # Find all files matching the pattern
        list_start = time.time()
        LOGGER.debug(T("coal.logs.dataset.listing_workspace_files"))
        all_api_files = api_ws.find_all_workspace_files(organization_id, workspace_id)
        
        existing_files = list(_f.file_name for _f in all_api_files if _f.file_name.startswith(file_name))
        list_time = time.time() - list_start
        
        LOGGER.debug(T("coal.logs.dataset.workspace_files_found").format(count=len(existing_files)))
        LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
            operation="file listing",
            time=list_time
        ))
        
        if not existing_files:
            LOGGER.warning(T("coal.logs.dataset.no_files_found").format(file_name=file_name))
            return content, Path(tmp_dataset_dir)
        
        # Download and process each file
        for _file_name in existing_files:
            download_start = time.time()
            LOGGER.debug(T("coal.logs.dataset.downloading_file").format(file_name=_file_name))
            
            dl_file = api_ws.download_workspace_file(
                organization_id=organization_id,
                workspace_id=workspace_id,
                file_name=_file_name
            )
            
            target_file = os.path.join(tmp_dataset_dir, _file_name.split('/')[-1])
            with open(target_file, "wb") as tmp_file:
                tmp_file.write(dl_file)
                
            download_time = time.time() - download_start
            LOGGER.debug(T("coal.logs.dataset.file_downloaded").format(
                file_name=_file_name,
                path=target_file
            ))
            LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
                operation=f"download {_file_name}",
                time=download_time
            ))
            
            if not read_files:
                continue
                
            # Process file based on type
            process_start = time.time()
            
            if ".xls" in _file_name:
                LOGGER.debug(T("coal.logs.dataset.processing_excel").format(file_name=target_file))
                wb = load_workbook(target_file, data_only=True)
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    content[sheet_name] = list()
                    headers = next(sheet.iter_rows(max_row=1, values_only=True))
                    
                    def item(_row: tuple) -> dict:
                        return {k: v for k, v in zip(headers, _row)}
                    
                    row_count = 0
                    for r in sheet.iter_rows(min_row=2, values_only=True):
                        row = item(r)
                        new_row = dict()
                        
                        for key, value in row.items():
                            try:
                                converted_value = json.load(io.StringIO(value))
                            except (json.decoder.JSONDecodeError, TypeError):
                                converted_value = value
                                
                            if converted_value is not None:
                                new_row[key] = converted_value
                                
                        if new_row:
                            content[sheet_name].append(new_row)
                            row_count += 1
                            
                    LOGGER.debug(T("coal.logs.dataset.sheet_processed").format(
                        sheet_name=sheet_name,
                        rows=row_count
                    ))
                    
            elif ".csv" in _file_name:
                LOGGER.debug(T("coal.logs.dataset.processing_csv").format(file_name=target_file))
                with open(target_file, "r") as file:
                    current_filename = os.path.basename(target_file)[:-len(".csv")]
                    content[current_filename] = list()
                    
                    row_count = 0
                    for csv_row in csv.DictReader(file):
                        csv_row: dict
                        new_row = dict()
                        
                        for key, value in csv_row.items():
                            try:
                                # Try to convert any json row to dict object
                                converted_value = json.load(io.StringIO(value))
                            except json.decoder.JSONDecodeError:
                                converted_value = value
                                
                            if converted_value == '':
                                converted_value = None
                                
                            if converted_value is not None:
                                new_row[key] = converted_value
                                
                        content[current_filename].append(new_row)
                        row_count += 1
                        
                    LOGGER.debug(T("coal.logs.dataset.csv_processed").format(
                        file_name=current_filename,
                        rows=row_count
                    ))
                    
            elif ".json" in _file_name:
                LOGGER.debug(T("coal.logs.dataset.processing_json").format(file_name=target_file))
                with open(target_file, "r") as _file:
                    current_filename = os.path.basename(target_file)
                    content[current_filename] = json.load(_file)
                    
                    if isinstance(content[current_filename], dict):
                        item_count = len(content[current_filename])
                    elif isinstance(content[current_filename], list):
                        item_count = len(content[current_filename])
                    else:
                        item_count = 1
                        
                    LOGGER.debug(T("coal.logs.dataset.json_processed").format(
                        file_name=current_filename,
                        items=item_count
                    ))
                    
            else:
                LOGGER.debug(T("coal.logs.dataset.processing_text").format(file_name=target_file))
                with open(target_file, "r") as _file:
                    current_filename = os.path.basename(target_file)
                    content[current_filename] = "\n".join(line for line in _file)
                    
                    line_count = content[current_filename].count('\n') + 1
                    LOGGER.debug(T("coal.logs.dataset.text_processed").format(
                        file_name=current_filename,
                        lines=line_count
                    ))
            
            process_time = time.time() - process_start
            LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
                operation=f"process {_file_name}",
                time=process_time
            ))
    
    elapsed_time = time.time() - start_time
    LOGGER.info(T("coal.logs.dataset.operation_timing").format(
        operation="File download",
        time=elapsed_time
    ))
    LOGGER.info(T("coal.logs.dataset.download_completed").format(dataset_type="File"))
    
    return content, Path(tmp_dataset_dir)


def download_dataset_by_id(
    organization_id: str,
    workspace_id: str,
    dataset_id: str,
    target_folder: Optional[Union[str, Path]] = None
) -> Tuple[Dict[str, Any], Path]:
    """
    Download dataset by ID.
    
    Args:
        organization_id: Organization ID
        workspace_id: Workspace ID
        dataset_id: Dataset ID
        target_folder: Optional folder to save files (if None, uses temp dir)
        
    Returns:
        Tuple of (dataset info dict, folder path)
    """
    start_time = time.time()
    LOGGER.info(T("coal.logs.dataset.download_started").format(dataset_type="Dataset"))
    LOGGER.debug(T("coal.logs.dataset.dataset_downloading").format(
        organization_id=organization_id,
        dataset_id=dataset_id
    ))
    
    with get_api_client()[0] as api_client:
        api_instance = DatasetApi(api_client)
        
        # Get dataset info
        info_start = time.time()
        dataset = api_instance.find_dataset_by_id(
            organization_id=organization_id,
            dataset_id=dataset_id
        )
        info_time = time.time() - info_start
        
        LOGGER.debug(T("coal.logs.dataset.dataset_info_retrieved").format(
            dataset_name=dataset.name,
            dataset_id=dataset_id
        ))
        LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
            operation="dataset info retrieval",
            time=info_time
        ))
        
        # Determine dataset type and download
        if dataset.connector is None:
            parameters = []
        else:
            parameters = dataset.connector.parameters_values
            
        is_adt = 'AZURE_DIGITAL_TWINS_URL' in parameters
        is_storage = 'AZURE_STORAGE_CONTAINER_BLOB_PREFIX' in parameters
        is_legacy_twin_cache = 'TWIN_CACHE_NAME' in parameters and dataset.twingraph_id is None
        is_in_workspace_file = False if dataset.tags is None else 'workspaceFile' in dataset.tags or 'dataset_part' in dataset.tags
        
        download_start = time.time()
        
        if is_adt:
            LOGGER.debug(T("coal.logs.dataset.dataset_type_detected").format(type="ADT"))
            content, folder = download_adt_dataset(
                adt_address=parameters['AZURE_DIGITAL_TWINS_URL'],
                target_folder=target_folder
            )
            dataset_type = 'adt'
            
        elif is_legacy_twin_cache:
            LOGGER.debug(T("coal.logs.dataset.dataset_type_detected").format(type="Legacy TwinGraph"))
            twin_cache_name = parameters['TWIN_CACHE_NAME']
            content, folder = download_legacy_twingraph_dataset(
                organization_id=organization_id,
                cache_name=twin_cache_name,
                target_folder=target_folder
            )
            dataset_type = 'twincache'
            
        elif is_storage or is_in_workspace_file:
            if is_storage:
                LOGGER.debug(T("coal.logs.dataset.dataset_type_detected").format(type="Storage"))
                _file_name = parameters['AZURE_STORAGE_CONTAINER_BLOB_PREFIX'].replace('%WORKSPACE_FILE%/', '')
            else:
                LOGGER.debug(T("coal.logs.dataset.dataset_type_detected").format(type="Workspace File"))
                _file_name = dataset.source.location
                
            content, folder = download_file_dataset(
                organization_id=organization_id,
                workspace_id=workspace_id,
                file_name=_file_name,
                target_folder=target_folder
            )
            dataset_type = _file_name.split('.')[-1]
            
        else:
            LOGGER.debug(T("coal.logs.dataset.dataset_type_detected").format(type="TwinGraph"))
            content, folder = download_twingraph_dataset(
                organization_id=organization_id,
                dataset_id=dataset_id,
                target_folder=target_folder
            )
            dataset_type = 'twincache'
            
        download_time = time.time() - download_start
        LOGGER.debug(T("coal.logs.dataset.operation_timing").format(
            operation="content download",
            time=download_time
        ))
    
    # Prepare result
    dataset_info = {
        "type": dataset_type,
        "content": content,
        "name": dataset.name
    }
    
    elapsed_time = time.time() - start_time
    LOGGER.info(T("coal.logs.dataset.operation_timing").format(
        operation="total dataset download",
        time=elapsed_time
    ))
    LOGGER.info(T("coal.logs.dataset.download_completed").format(dataset_type="Dataset"))
    
    return dataset_info, folder
