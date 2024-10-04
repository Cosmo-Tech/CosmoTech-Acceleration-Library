# Copyright (c) Cosmo Tech corporation.
# Licensed under the MIT license.
import csv
import io
import json
import multiprocessing
import os
import tempfile
from typing import Union

import cosmotech_api
from azure.digitaltwins.core import DigitalTwinsClient
from azure.identity import DefaultAzureCredential
from cosmotech_api import DatasetApi
from cosmotech_api import ScenarioApi
from cosmotech_api import TwingraphApi
from cosmotech_api import WorkspaceApi
from cosmotech_api import DatasetTwinGraphQuery
from cosmotech_api import TwinGraphQuery
from openpyxl import load_workbook

from cosmotech.coal.cosmotech_api.connection import get_api_client


def get_content_from_twin_graph_data(nodes, relationships, restore_names=False):
    '''
    When restore_names is True, the "id" value inside the "properties" field in the cypher query response is used
    instead of the numerical id found in the "id" field. When restore_names is set to False, this function
    keeps the previous behavior implemented when adding support for twingraph in v2 (default: False)

    Example with a sample of cypher response:
    [{
      n: {
        id: "50"  <-- this id is used if restore_names is False
        label: "Customer"
        properties: {
          Satisfaction: 0
          SurroundingSatisfaction: 0
          Thirsty: false
          id: "Lars_Coret"  <-- this id is used if restore_names is True
        }
        type: "NODE"
      }
    }]
    '''
    content = dict()
    # build keys
    for item in relationships:
        content[item['src']['label']] = list()
        content[item['dest']['label']] = list()
        content[item['rel']['label']] = list()

    for item in nodes:
        label = item['n']['label']
        props = item['n']['properties']
        if not restore_names:
            props.update({'id': item['n']['id']})
        content.setdefault(label, list())
        content[label].append(props)

    for item in relationships:
        src = item['src']
        dest = item['dest']
        rel = item['rel']
        props = item['rel']['properties']
        content[rel['label']].append({
            'id': rel['id'],
            'source': src['properties']['id'] if restore_names else src['id'],
            'target': dest['properties']['id'] if restore_names else dest['id'],
            **props
        })
    return content


class ScenarioDownloader:
    def __init__(
        self,
        workspace_id: str,
        organization_id: str,
        access_token: str = None,
        read_files=True,
        parallel=True
    ):
        if get_api_client()[1] == "Azure Entra Connection":
            self.credentials = DefaultAzureCredential()
        else:
            self.credentials = None
        

        self.workspace_id = workspace_id
        self.organization_id = organization_id
        self.dataset_file_temp_path = dict()
        self.read_files = read_files
        self.parallel = parallel

    def get_scenario_data(self, scenario_id: str):
        with get_api_client()[0] as api_client:
            api_instance = ScenarioApi(api_client)
            scenario_data = api_instance.find_scenario_by_id(organization_id=self.organization_id,
                                                             workspace_id=self.workspace_id,
                                                             scenario_id=scenario_id)
        return scenario_data

    def download_dataset(self, dataset_id: str) -> (str, str, Union[str, None]):
        with get_api_client()[0] as api_client:
            api_instance = DatasetApi(api_client)

            dataset = api_instance.find_dataset_by_id(
                organization_id=self.organization_id,
                dataset_id=dataset_id)
            parameters = dataset.connector.parameters_values

            is_adt = 'AZURE_DIGITAL_TWINS_URL' in parameters
            is_storage = 'AZURE_STORAGE_CONTAINER_BLOB_PREFIX' in parameters
            is_legacy_twin_cache = 'TWIN_CACHE_NAME' in parameters and dataset.twingraph_id is None  # Legacy twingraph dataset with specific connector

            if is_adt:
                return {
                    "type": 'adt',
                    "content": self._download_adt_content(
                        adt_adress=parameters['AZURE_DIGITAL_TWINS_URL']),
                    "name": dataset.name}
            elif is_legacy_twin_cache:
                twin_cache_name = parameters['TWIN_CACHE_NAME']
                return {
                    "type": "twincache",
                    "content": self._read_legacy_twingraph_content(twin_cache_name),
                    "name": dataset.name
                }
            elif is_storage:
                _file_name = parameters['AZURE_STORAGE_CONTAINER_BLOB_PREFIX'].replace(
                    '%WORKSPACE_FILE%/', '')
                _content = self._download_file(_file_name)
                self.dataset_file_temp_path[dataset_id] = self.dataset_file_temp_path[_file_name]
                return {
                    "type": _file_name.split('.')[-1],
                    "content": _content,
                    "name": dataset.name
                }
            else:
                return {
                    "type": "twincache",
                    "content": self._read_twingraph_content(dataset_id),
                    "name": dataset.name
                }

    def _read_twingraph_content(self, dataset_id: str) -> dict:
        with get_api_client()[0] as api_client:
            dataset_api = DatasetApi(api_client)
            nodes_query = DatasetTwinGraphQuery(query="MATCH(n) RETURN n")
            edges_query = DatasetTwinGraphQuery(query="MATCH(n)-[r]->(m) RETURN n as src, r as rel, m as dest")

            nodes = dataset_api.twingraph_query(
                organization_id=self.organization_id,
                dataset_id=dataset_id,
                dataset_twin_graph_query=nodes_query
            )
            edges = dataset_api.twingraph_query(
                organization_id=self.organization_id,
                dataset_id=dataset_id,
                dataset_twin_graph_query=edges_query
            )
            return get_content_from_twin_graph_data(nodes, edges, True)

    def _read_legacy_twingraph_content(self, cache_name: str) -> dict:
        with get_api_client()[0] as api_client:
            api_instance = TwingraphApi(api_client)
            _query_nodes = TwinGraphQuery(
                query="MATCH(n) RETURN n"
            )

            nodes = api_instance.query(
                organization_id=self.organization_id,
                graph_id=cache_name,
                twin_graph_query=_query_nodes
            )
            _query_rel = TwinGraphQuery(
                query="MATCH(n)-[r]->(m) RETURN n as src, r as rel, m as dest"
            )
            rel = api_instance.query(
                organization_id=self.organization_id,
                graph_id=cache_name,
                twin_graph_query=_query_rel
            )
            return get_content_from_twin_graph_data(nodes, rel, False)

    def _download_file(self, file_name: str):
        tmp_dataset_dir = tempfile.mkdtemp()
        self.dataset_file_temp_path[file_name] = tmp_dataset_dir
        with get_api_client()[0] as api_client:
            api_ws = WorkspaceApi(api_client)

            all_api_files = api_ws.find_all_workspace_files(
                self.organization_id, self.workspace_id)

            existing_files = list(
                _f.file_name for _f in all_api_files
                if _f.file_name.startswith(file_name))

            content = dict()

            for _file_name in existing_files:
                dl_file = api_ws.download_workspace_file(organization_id=self.organization_id,
                                                         workspace_id=self.workspace_id,
                                                         file_name=_file_name)

                target_file = os.path.join(
                    tmp_dataset_dir, _file_name.split('/')[-1])
                with open(target_file, "wb") as tmp_file:
                    tmp_file.write(dl_file)
                if not self.read_files:
                    continue
                if ".xls" in _file_name:
                    wb = load_workbook(target_file, data_only=True)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        content[sheet_name] = list()
                        headers = next(sheet.iter_rows(
                            max_row=1, values_only=True))

                        def item(_row: tuple) -> dict:
                            return {k: v for k, v in zip(headers, _row)}

                        for r in sheet.iter_rows(min_row=2, values_only=True):
                            row = item(r)
                            new_row = dict()
                            for key, value in row.items():
                                try:
                                    converted_value = json.load(
                                        io.StringIO(value))
                                except (json.decoder.JSONDecodeError, TypeError):
                                    converted_value = value
                                if converted_value is not None:
                                    new_row[key] = converted_value
                            if new_row:
                                content[sheet_name].append(new_row)
                elif ".csv" in _file_name:
                    with open(target_file, "r") as file:
                        # Read every file in the input folder
                        current_filename = os.path.basename(target_file)[:-len(".csv")]
                        content[current_filename] = list()
                        for csv_row in csv.DictReader(file):
                            csv_row: dict
                            new_row = dict()
                            for key, value in csv_row.items():
                                try:
                                    # Try to convert any json row to dict object
                                    converted_value = json.load(
                                        io.StringIO(value))
                                except json.decoder.JSONDecodeError:
                                    converted_value = value
                                if converted_value == '':
                                    converted_value = None
                                if converted_value is not None:
                                    new_row[key] = converted_value
                            content[current_filename].append(new_row)
                elif ".json" in _file_name:
                    with open(target_file, "r") as _file:
                        current_filename = os.path.basename(target_file)
                        content[current_filename] = json.load(_file)
                else:
                    with open(target_file, "r") as _file:
                        current_filename = os.path.basename(target_file)
                        content[current_filename] = "\n".join(
                            line for line in _file)
        return content

    def _download_adt_content(self, adt_adress: str) -> dict:
        client = DigitalTwinsClient(adt_adress, self.credentials)
        query_expression = 'SELECT * FROM digitaltwins'
        query_result = client.query_twins(query_expression)
        json_content = dict()
        for twin in query_result:
            entity_type = twin.get('$metadata').get(
                '$model').split(':')[-1].split(';')[0]
            t_content = {k: v for k, v in twin.items()}
            t_content['id'] = t_content['$dtId']
            for k in twin.keys():
                if k[0] == '$':
                    del t_content[k]
            json_content.setdefault(entity_type, [])
            json_content[entity_type].append(t_content)

        relations_query = 'SELECT * FROM relationships'
        query_result = client.query_twins(relations_query)
        for relation in query_result:
            tr = {
                "$relationshipId": "id",
                "$sourceId": "source",
                "$targetId": "target"
            }
            r_content = {k: v for k, v in relation.items()}
            for k, v in tr.items():
                r_content[v] = r_content[k]
            for k in relation.keys():
                if k[0] == '$':
                    del r_content[k]
            json_content.setdefault(relation['$relationshipName'], [])
            json_content[relation['$relationshipName']].append(r_content)

        return json_content

    def get_all_parameters(self, scenario_id) -> dict:
        scenario_data = self.get_scenario_data(scenario_id=scenario_id)
        content = dict()
        for parameter in scenario_data.parameters_values:
            content[parameter.parameter_id] = parameter.value
        return content

    def get_all_datasets(self, scenario_id: str) -> dict:
        scenario_data = self.get_scenario_data(scenario_id=scenario_id)

        datasets = scenario_data.dataset_list

        dataset_ids = datasets[:]

        for parameter in scenario_data.parameters_values:
            if parameter.var_type == '%DATASETID%':
                dataset_id = parameter.value
                dataset_ids.append(dataset_id)

        def download_dataset_process(_dataset_id, _return_dict, _error_dict):
            try:
                _c = self.download_dataset(_dataset_id)
                if _dataset_id in self.dataset_file_temp_path:
                    _return_dict[_dataset_id] = (_c, self.dataset_file_temp_path[_dataset_id], _dataset_id)
                else:
                    _return_dict[_dataset_id] = _c
            except Exception as e:
                _error_dict[_dataset_id] = f'{type(e).__name__}: {str(e)}'
                raise e

        if self.parallel:
            manager = multiprocessing.Manager()
            return_dict = manager.dict()
            error_dict = manager.dict()
            processes = [
                (dataset_id, multiprocessing.Process(target=download_dataset_process,
                                                     args=(dataset_id, return_dict, error_dict)))
                for dataset_id in dataset_ids
            ]
            [p.start() for _, p in processes]
            [p.join() for _, p in processes]

            for dataset_id, p in processes:
                # We might hit the following bug: https://bugs.python.org/issue43944
                # As a workaround, only treat non-null exit code as a real issue if we also have stored an error
                # message
                if p.exitcode != 0 and dataset_id in error_dict:
                    raise ChildProcessError(
                        f"Failed to download dataset '{dataset_id}': {error_dict[dataset_id]}")
        else:
            return_dict = {}
            error_dict = {}
            for dataset_id in dataset_ids:
                try:
                    download_dataset_process(dataset_id, return_dict, error_dict)
                except Exception as e:
                    raise ChildProcessError(
                        f"Failed to download dataset '{dataset_id}': {error_dict.get(dataset_id, '')}")
        content = dict()
        for k, v in return_dict.items():
            if isinstance(v, tuple):
                content[k] = v[0]
                self.dataset_file_temp_path[v[2]] = v[1]
            else:
                content[k] = v
        return content

    def dataset_to_file(self, dataset_id, dataset_info):
        type = dataset_info['type']
        content = dataset_info['content']
        name = dataset_info['name']
        if type in ["adt", "twincache"]:
            return self.adt_dataset(content, name, type)
        return self.dataset_file_temp_path[dataset_id]

    @staticmethod
    def sheet_to_header(sheet_content):
        fieldnames = []
        has_src = False
        has_id = False
        for r in sheet_content:
            for k in r.keys():
                if k not in fieldnames:
                    if k in ['source', 'target']:
                        has_src = True
                    elif k == "id":
                        has_id = True
                    else:
                        fieldnames.append(k)
        if has_src:
            fieldnames = ['source', 'target'] + fieldnames
        if has_id:
            fieldnames = ['id', ] + fieldnames
        return fieldnames

    def adt_dataset(self, content, _name, _type):
        tmp_dataset_dir = tempfile.mkdtemp()
        for _filename, _filecontent in content.items():
            with open(tmp_dataset_dir + "/" + _filename + ".csv", "w") as _file:
                fieldnames = self.sheet_to_header(_filecontent)

                _w = csv.DictWriter(_file, fieldnames=fieldnames, dialect="unix", quoting=csv.QUOTE_MINIMAL)
                _w.writeheader()
                # _w.writerows(_filecontent)
                for r in _filecontent:
                    _w.writerow(
                        {k: str(v).replace("'", "\"").replace("True", "true").replace("False", "false") for k, v in
                         r.items()})
        return tmp_dataset_dir
